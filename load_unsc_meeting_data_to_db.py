from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy.exc import IntegrityError

import argparse
import concurrent.futures
import fitz
import logging
import openpyxl
import os
import pickle
import re
import threading

from html_downloader import HTMLDownloader
from meeting_html_parser import MeetingHTMLParser
from job_queue import JobQueue
from job import Job
from meeting_job import MeetingJob
from pdf_downloader import PDFDownloader
from veto_html_parser import VetoHTMLParser
from vetocasts import VetoCasts
from meeting import Meeting
from resolution import Resolution
from state import State
from dbconnection import DBConnection

logger = logging.getLogger("unsc_db_filler")
sql_logger = logging.getLogger("sqlalchemy")
logger.setLevel(logging.DEBUG)
sql_logger.setLevel(logging.INFO)

ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
logger.addHandler(ch)
sql_logger.addHandler(ch)

fh = logging.FileHandler(f"Logs/logging-{datetime.now().isoformat()}.log")
fh_formatter = logging.Formatter(
    "%(asctime)s - %(levelname)s - %(threadName)s: %(message)s"
)
fh.setFormatter(fh_formatter)
fh.setLevel(logging.INFO)
logger.addHandler(fh)
sql_logger.addHandler(fh)

# Read the DB info and credentials from our .env file
load_dotenv("Database/database.env")
DB_HOSTNAME = os.getenv("POSTGRES_HOSTNAME")
DB_USERNAME = os.getenv("POSTGRES_USER")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD")
DB_NAME = os.getenv("POSTGRES_DB")

# Folder to download UNSC Resolution PDFs to
SCRATCH_FOLDER = "UNDataScraping/scratch"
RESOLUTION_DOWNLOAD_FOLDER = f"{SCRATCH_FOLDER}/unsc_resolution_pdfs/"
MEETING_DOWNLOAD_FOLDER = f"{SCRATCH_FOLDER}/unsc_meeting_pdfs/"
UN_LIBRARY_FILE1 = "UNExports/089_B02_-_SECURITY_COUNCIL_-_1st_Link.xlsx"
# A global variable to store the VETO Table in
VETO_TABLE = None

# DB Connection
db_connection = DBConnection(
    host=DB_HOSTNAME, dbname=DB_NAME, user=DB_USERNAME, password=DB_PASSWORD
)


class DownloadFailed(Exception):
    pass


def initialize_state_table():
    """
    This function is used to intialize the state table in the DB.
    The state table stores the permanent UNSC members. If ever additional P5 members are added,
    this function needs to be updated.

    :return:
    """
    session = db_connection.get_session()

    p5_member_states = ["USA", "UK", "France", "China", "Russia"]

    for state in p5_member_states:
        session.add(State(name=state))

    try:
        session.commit()
    except IntegrityError:
        logger.info("Record already exists")
        session.rollback()


def read_from_scratch(filename: str) -> str:
    """
    Reads the UNSC meeting html file in the scratch directory.

    :param filename: Which downloaded UNSC Meeting page is to be read
    :return: The raw HTML of the UNSC meeting table
    """
    input_file = f"{SCRATCH_FOLDER}/{filename}"
    logger.info("Reading content from file '%s' BEGIN", input_file)

    with open(input_file, "r") as f:
        html = f.read()
        logger.info("Reading content from file '%s' END", input_file)

    return html


def read_pdf(pdf: str) -> str:
    """
    Reads the text of a given pdf file, and returns it as unicode UTF-8 encoded
    text
    :param pdf: The Path to the PDF
    :return:    The text found inside the PDF UTF-8 encoded
    """
    with fitz.open(f"{pdf}") as doc:
        text = ""
        for page in doc:
            text += page.get_text()

    return text


def find_not_adopted_draft_resolution_mentioned_in(target: str) -> list:
    """
    Checks if a string has draft resolutions mentioned.
    A draft resolution has the format `S/YYYY/X` - `S/YYYY/XXXX`
    where YYYY is the year and X is the Xth draft resolution of that year.
    In the earlier days of the UNSC, draft resolutions had the format `S/X - S/XXXXX`
    If it finds any, it returns them as a list.

    :param target: The string we're searching in
    :return: a list of found draft resolutions
    """
    return re.findall(
        "(S/[0-9]{1,5}|S/PV.[0-9]{1,4}|S/[0-9]{4}/[0-9]{1,4})\s*\n*\t*\r*not adopted", target
    )


def find_vetoed_draft_resolution_mentioned_in(target: str) -> list:
    """
    Checks if a string has vetoed draft resolutions mentioned.
    A vetoed draft resolution has the format `S/YYYY/X vetoed by` - `S/YYYY/XXXX vetoed by`
    where YYYY is the year and X is the Xth draft resolution of that year.
    In the earlier days of the UNSC, draft resolutions had the format `S/X - S/XXXXX`
    In the first 2 years of the UNSC, draft resolutions also show up as `S/PV.X`
    In 1986 we also see draft resolutions of the form `S/XXXX/Rev.Y`

    If it finds any, it returns them as a list.

    :param target: The string we're searching in
    :return: a list of found vetoed draft resolutions
    """
    return re.findall(
        #"\(?(S/[0-9]{1,5}/Rev.[0-9]+|S/PV.[0-9]{1,4}|S/[0-9]{1,5}|S/[0-9]{4}/[0-9]{1,4})\)?\s*\n*\t*\r*vetoed by",
        #"\(?(S/[0-9]{1,5}/Rev.[0-9]+|S/PV.[0-9]{1,4}|S/[0-9]{1,5}|S/[0-9]{4}/[0-9]{1,4})\)?\s*\n*\t*\r*\(?[a-zA-Z]*\)*\s*\n*\t*\r*vetoed by",
        "\(?(S/[0-9]{1,5}/Rev.[0-9]+|S/PV.[0-9]{1,4}|S/[0-9]{1,5}|S/[0-9]{4}/[0-9]{1,4})\)?\s*\n*\t*\r*\(?[a-zA-Z\s]*\)*\s*\n*\t*\r*vetoed by",
        target,
    )


def find_vetoed_draft_resolution_mentioned_in_by_partial_match(vetoed_res: str) -> str:
    """
    Checks the VETO_TABLE for a partial match..if we find one, return that match
    Normally not required to be called. Only when the normal flow did not find the resolution
    in the VETO_TABLE, yet still we believe it's there.

    For example:
    S/PV.2686 discussed draft resolution S/18087 according to https://www.un.org/depts/dhl/resguide/scact1986_table_en.html
    However, the veto table mentions S/18087.Rev1. S/18087 won't be found using an exact match.
    If we do a partial search, we can find it.

    :param vetoed_res: The vetoed resolution we're searching the vetoed_by person for
    :return: the draft resolution as listed in the VETO_TABLE
    """
    for key, value in VETO_TABLE.items():
        if re.match(vetoed_res, key):
            return key

    # We shouldn't get here... if we would, it means there's some serious discrepancy between
    # what's on the UNSC Meeting pages and on the veto table.
    return ""


def find_adopted_resolution_mentioned_in(target: str) -> list:
    """
    Checks if a string has adopted resolutions mentioned.
    Format of an approved Resolution is: `S/RES/X` - `S/RE/XXXX (YYYY)` where X is the number
    and YYYY is the year

    If it finds any, it returns them as a list.

    :param target: The string we're searching in
    :return: a list of found adopted resolutions
    """

    # Some resolutions have a space in their name.
    # Which is not how they are referenced in URLs and other data sources.
    # Because of that, we remove those spaces first... .
    target = re.sub("\s+", "", target)
    return re.findall("S\/RES\/[0-9]{1,4}\s?\(?[0-9]{4}\)?", target)


def find_draft_resolution_for_adopted_resolution(resolution: str) -> str:
    """
    Given an accepted resolution `S/RES/XXXXX`, what is the original draft resolution?
    Note, this requires a data export from the UN Library. See README.md, you need `file1.xlsx`

    :param resolution: the final accepted resolution ID
    :return: the original draft resolution ID
    """
    logger.info("Finding draft for adopted resolution '%s'", resolution)
    book = openpyxl.load_workbook(UN_LIBRARY_FILE1)
    sheet = book["results"]
    for row in range(1, sheet.max_row + 1):

        # At the moment the adopted resolution is at column J,
        # and the original one at column B.
        # However..you never know what they might change later again,
        # so let's just search across 26 columns at least... .
        for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            cell = f"{column}{row}"
            if sheet[cell].value == resolution:
                # We found the cell of the adopted resolution..
                # On the same row, but in column B, we have the
                # original resolution
                orig_cell = f"B{row}"
                return sheet[orig_cell].value

    # If we got here, it means the adopted resolution wasn't found,
    # or the draft resolution is missing. More than usual, the adopted resolution
    # wasn't found through exact matching. We've seen data inconsistencies there.
    # Example S/RES/2498(2019) shows in the Excel sheet as S/RES/2498(2018).
    # The year is wrong! Let's try to get rid of that suffix (2018) and do some partial matching...
    # TODO: Fix duplication
    resolution = re.sub("\(.*\)", "", resolution)
    logger.info("Searching for '%s(' in the Excel sheet (partial matching)...", resolution)
    for row in range(1, sheet.max_row + 1):

        # At the moment the adopted resolution is at column J,
        # and the original one at column B.
        # However..you never know what they might change later again,
        # so let's just search across 26 columns at least... .
        for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            cell = f"{column}{row}"
            if re.match(f"^{resolution}\(", str(sheet[cell].value)):
                # We found the cell of the adopted resolution..
                # On the same row, but in column B, we have the
                # original resolution
                orig_cell = f"B{row}"
                return sheet[orig_cell].value

    # If we got here, something is wrong with our original data...
    # We failed to retrieve the original resolution..
    return "UNKNOWN"


def process_job(job: MeetingJob) -> None:
    """
    This function is the heavy lifter of the data set building
    It is used as the process method on the Job Queue where all the Meeting Jobs are queued.

    This function creates the real objects (Meeting, Resolution, ...) that will be persisted to the database.


    :param job:
    :return:
    """

    from time import sleep
    from math import floor
    from random import random

    sleep(floor(random() * 10))

    db_session = db_connection.get_session()
    db_session.autoflush = False

    logger.info("Running in thread #%s'", threading.current_thread().name)

    meeting_record = job.meeting_record.get("meeting_record")

    year = job.meeting_record.get("year")
    date = job.meeting_record.get("date")
    outcome = job.meeting_record.get("outcome")
    topic = job.meeting_record.get("topic")
    meeting_url = job.meeting_record.get("meeting_url")

    try:
        meeting = Meeting(
            meeting_id=meeting_record,
            year=year,
            date=date,
            topic=topic,
            meeting_url=meeting_url,
        )

        meeting_pdf_downloader = PDFDownloader(path=f"{MEETING_DOWNLOAD_FOLDER}{year}")

        resolution_pdf_downloader = PDFDownloader(
            path=f"{RESOLUTION_DOWNLOAD_FOLDER}{year}"
        )

        logger.info("Downloading PDF for meeting '%s' - BEGIN", meeting_record)
        meeting_pdf = meeting_pdf_downloader.download_pdf(meeting_record, meeting_url)
        logger.info("Downloading PDF for meeting '%s' - END", meeting_record)

        logger.info("Extracting text from PDF for meeting '%s' - BEGIN", meeting_record)
        meeting_transcript = read_pdf(meeting_pdf)
        logger.info("Adding text to meeting '%s' - END", meeting_record)
        meeting.add_meeting_transcript(meeting_transcript)

        vetoed_draft_resolutions = find_vetoed_draft_resolution_mentioned_in(outcome)
        not_adopted_resolutions = find_not_adopted_draft_resolution_mentioned_in(
            outcome
        )
        adopted_resolutions = find_adopted_resolution_mentioned_in(outcome)

        # Iterate first over the vetoed Resolutions
        for vetoed_res in vetoed_draft_resolutions:

            # Data inconsistency...
            # S/PV.2686 discussed draft resolution S/18087 according to https://www.un.org/depts/dhl/resguide/scact1986_table_en.html
            # However, the veto table mentions S/18087.Rev1.
            # So we'll do a partial match... .
            if VETO_TABLE.get(vetoed_res) is None:
                vetoed_res = find_vetoed_draft_resolution_mentioned_in_by_partial_match(
                    vetoed_res
                )

            # Because of data inconsistency issues (we have seen that a resolution can be listed as vetoed on the
            # meeting overview by year webpages, but not on the veto table. If that's the case,
            # let's not crash but continue... .
            if vetoed_res is None:
                continue

            veto_voters = VETO_TABLE[vetoed_res]

            res = Resolution(
                draft_id=vetoed_res,
                final_id=None,
                kind=Resolution.DRAFT,
                adopted=False,
                vetoed=True,
                # vetoed_by=veto_voters,
                meeting=meeting,
                year=year,
            )

            logger.info("Downloading PDF for Resolution '%s' - BEGIN", vetoed_res)
            resolution_pdf = resolution_pdf_downloader.download_pdf(vetoed_res)
            logger.info("Downloading PDF for Resolution '%s' - END", vetoed_res)

            logger.info("Reading text from PDF '%s' - BEGIN", resolution_pdf)
            draft_text = read_pdf(resolution_pdf)
            logger.info("Reading text from PDF '%s' - END", resolution_pdf)
            res.add_draft_text(draft_text)

            db_session.merge(res)

            for veto_voter in veto_voters:
                vetoing_state = (
                    db_session.query(State).filter(State.name == veto_voter).first()
                )
                logger.info("Veto voter: %s", veto_voter)
                logger.info("Veto voter id: %s", vetoing_state.state_id)
                db_session.add(VetoCasts(vetoed_res, vetoing_state.state_id))

            db_session.commit()

        # Iterate over the not adopted Resolutions
        for not_adopted_res in not_adopted_resolutions:
            res = Resolution(
                draft_id=not_adopted_res,
                final_id=None,
                kind=Resolution.DRAFT,
                adopted=False,
                vetoed=False,
                # vetoed_by=None,
                meeting=meeting,
                year=year,
            )

            logger.info("Downloading PDF for Resolution '%s' - BEGIN", not_adopted_res)

            resolution_pdf = resolution_pdf_downloader.download_pdf(not_adopted_res)
            logger.info("Downloading PDF for Resolution '%s' - END", not_adopted_res)

            logger.info("Reading text from PDF '%s' - BEGIN", resolution_pdf)
            draft_text = read_pdf(resolution_pdf)
            logger.info("Reading text from PDF '%s' - END", resolution_pdf)
            res.add_draft_text(draft_text)

            db_session.merge(res)
            db_session.commit()

        # Iterate over the adopted Resolutions
        for adopted_res in adopted_resolutions:
            logger.info(
                f"Searching for draft resolution of resolution '%s' in UN Library Excel sheet",
                adopted_res,
            )
            draft_res = find_draft_resolution_for_adopted_resolution(adopted_res)
            if draft_res == "UNKNOWN":
                raise Exception(
                    f"Failed to find the draft resolution for resolution '{adopted_res}' ({meeting_record}) in the Excel sheet"
                )

                # We assume that if it is not listed in the Excel sheet,
                # it is a very old resolution where there were not at all times
                # draft resolutions for every discussion.
                #draft_res = adopted_res
            logger.info(
                f"Draft resolution of resolution '%s' found in UN Library Excel sheet: %s",
                adopted_res,
                draft_res,
            )

            res = Resolution(
                draft_id=draft_res,
                final_id=adopted_res,
                kind=Resolution.FINAL,
                adopted=True,
                vetoed=False,
                # vetoed_by=None,
                meeting=meeting,
                year=year,
            )

            logger.info("Downloading PDF for Resolution '%s' - BEGIN", draft_res)

            draft_resolution_pdf = resolution_pdf_downloader.download_pdf(draft_res)
            logger.info("Downloading PDF for Resolution '%s' - END", draft_res)

            logger.info("Downloading PDF for Resolution '%s' - BEGIN", adopted_res)
            adopted_resolution_pdf = resolution_pdf_downloader.download_pdf(adopted_res)
            logger.info("Downloading PDF for Resolution '%s' - END", adopted_res)

            logger.info("Reading text from PDF '%s' - BEGIN", draft_resolution_pdf)
            draft_text = read_pdf(draft_resolution_pdf)
            logger.info("Reading text from PDF '%s' - END", draft_resolution_pdf)

            logger.info("Reading text from PDF '%s' - BEGIN", adopted_resolution_pdf)
            adopted_text = read_pdf(adopted_resolution_pdf)
            logger.info("Reading text from PDF '%s' - END", adopted_resolution_pdf)

            res.add_draft_text(draft_text)
            res.add_final_text(adopted_text)

            db_session.merge(res)
            db_session.commit()

    except IntegrityError as ie:
        logger.info("Error committing: %s", ie)
        logger.info("Record already exists in DB...skipping")
        db_session.rollback()
    finally:
        db_session.close()


def main() -> None:
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--since",
        help="Which year to start fetching meetings and resolutions from",
        action="store",
        type=int,
        default=1946,
    )

    parser.add_argument(
        "--until",
        help="Until (and including) which year to fetch meetings and resolutions from",
        action="store",
        type=int,
        default=2021,
    )

    parser.add_argument(
        "--fetch-all-unsc-tables",
        help="Forces a re-download of all the UNSC meeting table pages",
        action="store_true",
        default=False,
    )

    parser.add_argument(
        "--retry-file",
        help="Retry processing records from a previous run its failed file",
        action="store",
        type=str,
    )

    parser.add_argument(
        "--workers",
        help="The amount of workers you want to use for parallel processing",
        action="store",
        type=int,
        default=8,
    )

    if not os.path.exists(f"{SCRATCH_FOLDER}"):
        os.makedirs(f"{SCRATCH_FOLDER}")

    args = parser.parse_args()

    # Prepare our job processing queue
    job_queue = JobQueue()

    if args.fetch_all_unsc_tables:
        downloader = HTMLDownloader(
            path=SCRATCH_FOLDER, since=args.since, until=args.until
        )
        downloader.fetch_meeting_tables()
        downloader.fetch_veto_table()

        # The download queue should be empty.. if it's not we miss critical information
        # and should NOT continue.. we'll stop here.
        if len(downloader.job_queue.failed) > 0:
            logger.error(
                "!!! Not all UNSC tables were fetched.. Please retry at a later time or verify the un.org website did not change URLs. !!!"
            )
            exit(-1)

    # Prepare the Veto Mapping Dictionary
    html = read_from_scratch("scact_veto_table_en.html")
    veto_parser = VetoHTMLParser(html)
    global VETO_TABLE
    VETO_TABLE = veto_parser.create_resolution_to_veto_mapping_table()

    initialize_state_table()

    def fill_queue(from_year: int, end_year: int):
        for year in range(from_year, end_year):
            try:
                html = read_from_scratch(f"scact{year}_table_en.html")

                meeting_parser = MeetingHTMLParser(html)
                records: list = meeting_parser.extract_records()

                for record in records:
                    # Add the current year to the record
                    record["year"] = year

                    # Debug specific meeting record
                    #if record["meeting_record"].strip() != "S/PV.423":
                    #    continue

                    # Make jobs from each record, and queue it.
                    j = MeetingJob(record)
                    job_queue.enqueue(j)

            except Exception as e:
                logger.error("Failed reading html pages: %s", e)

            try:
                # Check for covid tables... if there's any, parse them..else, skip
                covid_table = f"scact{year}_covid_table_en.html"
                if not os.path.exists(f"{SCRATCH_FOLDER}/{covid_table}"):
                    continue
                else:
                    html = read_from_scratch(covid_table)
                    meeting_parser = MeetingHTMLParser(html)
                    records: list = meeting_parser.extract_records()

                    for record in records:
                        # Add the current year to the record
                        record["year"] = year

                        # Debug specific meeting record
                        #if record["meeting_record"].strip() != "S/PV.423":
                        #    continue

                        # Make jobs from each record, and queue it.
                        j = MeetingJob(record)
                        job_queue.enqueue(j)

            except Exception as e:
                logger.error("Failed reading covid html pages: %s", e)

        logger.info("%s jobs queued, ready for processing", job_queue.size())

    # If a user specified a --retry-file,
    # Unpickle and enqueue the jobs that previously failed and we want to retry now.
    if args.retry_file:
        logger.info("Retrying records in file '%s'", args.retry_file)
        failed_jobs = pickle.load(open(args.retry_file, "rb"))
        logger.info("%s jobs found in the file", len(failed_jobs))

        for job in failed_jobs:
            logger.info(job.meeting_record)

            # Debug specific meeting record
            # if job.meeting_record.get('meeting_id') != 'S/PV.2970':
            #    continue

            j = MeetingJob(job.meeting_record)
            job_queue.enqueue(j)

    # Else..if we do not specify a retry-file, just enqueue records between the --since and --until years.
    else:
        fill_queue(from_year=args.since, end_year=args.until)

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
        executor.submit(job_queue.process, process_job)
        executor.submit(job_queue.process, process_job)
        executor.submit(job_queue.process, process_job)

    logger.info("%s jobs remaining unprocessed", job_queue.size())
    logger.info("%s jobs failed to process", len(job_queue.failed))
    logger.info("Failed jobs: %s", job_queue.failed)

    # Save the failed jobs to a file for retry later
    if len(job_queue.failed) > 0:
        pickle.dump(
            job_queue.failed,
            open(f"failed_records-{datetime.now().isoformat()}.p", "wb"),
        )

        with open(f"failed_records-{datetime.now().isoformat()}.txt", "a") as f:
            for record in job_queue.failed:
                f.write(f"{record.meeting_record}\n")


if __name__ == "__main__":
    main()
